import operator
from datetime import datetime, timedelta
from aiogram.types import Message, FSInputFile
from openpyxl.styles import Border, Side
from utils.functions.get_bot_and_db import get_bot_and_db
from aiogram.exceptions import TelegramBadRequest
from blanks.bot_markup import choice_engines_status, choice_status_task, choice_start, choice_engine, choice_type_tasks
import asyncio
from openpyxl import Workbook, styles

bot, db, db_cpu = get_bot_and_db()


async def return_status(percent) -> str:
    last_value = percent[0]['Последнее_значение']
    last_value_rounded = round(last_value)
    status = ''
    if last_value_rounded < 20:
        status = "low 😴"
    elif 20 <= last_value <= 80:
        status = "normal 🆗"
    elif 80 < last_value <= 100:
        status = "high ❗"
    return status


async def check_servers_and_notify():
    chat_ids = [633496059,
                733619683,  # Игорь
                # 438576314  # Вадим
                ]
    statuses = dict()
    # Выполним проверку каждого сервера и формирование сообщения
    formatted_string = ""
    server_names = ['qlik01', 'qlik02', 'qlik03']
    current_time = datetime.now().strftime('%H:%M')
    current_hour = int(current_time[:2])

    for server_name in server_names:
        record = await db_cpu.return_cpu_or_memory(server_name, 1, 'CPU')
        last_value = record[0]['Последнее_значение']
        last_value_rounded = round(last_value)
        statuses[server_name] = last_value
        status = await return_status(record)
        formatted_string += f"{server_name} <b>{last_value_rounded}% - {status}</b>\n\n"

    # Отправляем уведомление всем пользователям если была нагрузка более 90%
    for engine, load in statuses.items():
        async def send_notify(limit):
            if load >= limit:
                text = f"Высокая нагрузка на сервер {engine} ❗\n\n" + formatted_string
                for chat_id in chat_ids:
                    await bot.send_message(chat_id, text=text, parse_mode='html')

        if current_hour >= 22 or current_hour < 8:  # Если время с 22:00 до 08:00
            if engine == 'qlik01':  # Если первый сервер
                await send_notify(98)  # отправляем сообщение только при нагрузке от 98 процентов
            else:
                await send_notify(90)  # если в это время нагрузка на другие сервера от 90 процентов
        else:  # В течение остального дня
            await send_notify(90)


async def text_handler(message: Message):
    tg_id = message.from_user.id
    m_id = message.message_id
    chat_type = message.chat.type
    text = message.text
    print(tg_id)

    async def write_to_file_and_send(filename: str, status=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks Table"

        headers = ["App Name", "Duration", "Status", "Start Time"]
        if not status:
            headers.append("Who Started")

        for col_num, header in enumerate(headers, start=1):
            col_letter = ws.cell(row=1, column=col_num).column_letter
            ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num)
            ws.cell(row=1, column=col_num).value = header
            # Выделение заголовков
            ws.cell(row=1, column=col_num).font = styles.Font(bold=True)
            ws.cell(row=1, column=col_num).fill = styles.PatternFill(start_color="ebe6ca", end_color="ebe6ca",
                                                                     fill_type="solid")
            ws.column_dimensions[col_letter].width = 20

        if status:  # Если был передан статус (2, 6, 7, 8), то таски по статусу
            tasks = await db.return_tasks_with_status(status)
        else:  # вручную перезапущенные таски без конкретного статуса
            tasks = await db.return_manually_tasks()

        tasks_sorted_by_time = sorted(tasks, key=operator.itemgetter('StartTime'), reverse=True)

        status_name = {
            2: "Started", 6: "Aborted", 7: "Success", 8: "Failed"
        }

        # Определение максимальной ширины для каждого столбца
        max_widths = {}
        for i, record in enumerate(tasks_sorted_by_time, start=2):
            columns = [record['app_name'], record['duration_seconds'], status_name[record['Status']],
                       record['StartTime'].strftime('%Y-%m-%d %H:%M:%S')]
            if not status:
                columns.append(record['ModifiedByUserName'])
            for j, value in enumerate(columns, start=1):
                cell_value = str(value)
                if len(cell_value) > max_widths.get(j, 0):
                    max_widths[j] = len(cell_value)

        # Настройка ширины столбцов
        for col_num, width in max_widths.items():
            col_letter = ws.cell(row=1, column=col_num).column_letter
            ws.column_dimensions[col_letter].width = width + 2  # Добавляем немного запаса

        # Границы для всех ячеек
        border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

        for i, record in enumerate(tasks_sorted_by_time, start=2):
            formatted_time = record['StartTime'].strftime('%Y-%m-%d %H:%M:%S')
            time_obj = datetime.strptime(formatted_time, '%Y-%m-%d %H:%M:%S')
            new_time_obj = time_obj + timedelta(hours=3)  # Добавление 3 часов
            formatted_time = new_time_obj.strftime('%Y-%m-%d %H:%M:%S')
            duration_seconds = record['duration_seconds'] - 10800  # 10800с = 3часа (разница из-за UTC)
            hours, remainder = divmod(duration_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            formatted_duration = f"{str(int(hours)).zfill(2)}:{str(int(minutes)).zfill(2)}:{str(int(seconds)).zfill(2)}"
            row = [record['app_name'], formatted_duration, status_name[record['Status']], formatted_time]
            if not status:
                row.append(record['ModifiedByUserName'])
            for j, cell_value in enumerate(row, start=1):
                ws.cell(row=i, column=j).value = cell_value

        filename = filename.replace('.txt', '.xlsx')
        wb.save(filename)

        document = FSInputFile(filename)
        await message.answer_document(document=document)

    # if text == 'Состояние сервера ⚙':
    #     await message.answer("Выберите посмотреть состояние серверов в QMS или нагрузку по CPU/RAM:",
    #                          reply_markup=choice_engines_status())

    if text == 'Контроль тасков 📋':
        await message.answer("Выберите нужный тип тасков:\n"
                             "<b>Manually</b> - запущены вручную/из хаба\n"
                             "<b>Scheduler</b> - по задаче обновления:",
                             reply_markup=choice_type_tasks(), parse_mode='html')

    elif text == 'Scheduler ⏳':
        await message.answer("Выберите интересующий статус тасков:",
                             reply_markup=choice_status_task())

    elif text == 'Manually 🔧':
        await message.answer("Таски, которые запущены вручную из хаба:")
        await write_to_file_and_send('manually_tasks.txt')

    # elif text == 'Состояние CPU/RAM 📈':
    elif text == 'Состояние сервера ⚙':
        await bot.send_chat_action(message.chat.id, action="typing", request_timeout=5)

        # qlik01 ttk
        record_central = await db_cpu.return_cpu_or_memory('qlik01', 1, 'CPU')
        last_value = record_central[0]['Последнее_значение']
        last_value_rounded = round(last_value)
        formatted_string = f"qlik01 (ttk) <b>{last_value_rounded}% - {await return_status(record_central)}</b>"
        await bot.send_chat_action(message.chat.id, action="typing", request_timeout=5)

        # qlik02 dev
        record_dev = await db_cpu.return_cpu_or_memory('qlik02', 1, 'CPU')
        last_value = record_dev[0]['Последнее_значение']
        last_value_rounded = round(last_value)
        formatted_string += f"\n\nqlik02 (dev) <b>{last_value_rounded}% - {await return_status(record_dev)}</b>"
        await bot.send_chat_action(message.chat.id, action="typing", request_timeout=5)

        # qlik03 pl
        record_qlik03 = await db_cpu.return_cpu_or_memory('qlik03', 1, 'CPU')
        last_value = record_qlik03[0]['Последнее_значение']
        last_value_rounded = round(last_value)
        formatted_string += (f"\n\nqlik03 (pl) <b>{last_value_rounded}% - {await return_status(record_qlik03)}</b>\n\n"
                             f"Выберите сервер для детальной информации:")
        await bot.send_chat_action(message.chat.id, action="typing", request_timeout=5)

        await message.answer(text=formatted_string, parse_mode='html', reply_markup=choice_engine().as_markup())

    elif text == 'Started 🔁':
        try:
            await message.answer(text=await db.get_tasks(2), parse_mode='html')
        except TelegramBadRequest as tbr:
            if not await db.get_tasks(2):
                await message.answer(text='Запущенных тасков нет 🎈', parse_mode='html')
            else:
                await write_to_file_and_send('started_tasks.txt', 2)

    elif text == 'Aborted ❎':
        try:
            await message.answer(text=await db.get_tasks(6), parse_mode='html')
        except TelegramBadRequest:
            await write_to_file_and_send('aborted_tasks.txt', 6)

    elif text == 'Failed ❌':
        try:
            await message.answer(text=await db.get_tasks(8), parse_mode='html')
        except TelegramBadRequest:
            await write_to_file_and_send('failed_tasks.txt', 8)

    elif text == 'Success ✔':
        await write_to_file_and_send('success_tasks.txt', 7)

    elif text == 'Назад 🔙':
        await message.answer(
            text="Здесь можно выбрать:\n"
                 "<b>Состояние сервера</b> - нагрузка на сервера по CPU/RAM\n"
                 "<b>Контроль тасков</b> - информация по таскам:",
            reply_markup=choice_start(),
            parse_mode='html')

    elif text == 'Назад ⏪':
        await message.answer("Выберите нужный тип тасков:\n"
                             "<b>Manually</b> - запущены вручную/из хаба\n"
                             "<b>Scheduler</b> - по задаче обновления:",
                             reply_markup=choice_type_tasks(), parse_mode='html')
